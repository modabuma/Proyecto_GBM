from sqlalchemy.orm import sessionmaker
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from io import BytesIO
import base64
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill

from models.UsersModel import UsersModel
from models.ObservabilityModel import ObservabilityModel
from shared.custom_exceptions import BadCredentialsError, NotFoundError

class AuthController:
    def __init__(self, session: sessionmaker):
        self.session = session
        
    def authenticate(self, payload: dict) -> dict:
        user_model = UsersModel()
        
        response = user_model.get_data(self.session, {
            "email": payload["email"], "role": 1,
            "active": 1
        })
        
        if not response:
            raise BadCredentialsError("El correo ingresado no coincide.")

        if not response.verify_password(payload["password"]):
            raise BadCredentialsError("La contraseña ingresada no coincide.")
            
        
        return {
            "id": response.id,
            "names": response.name,
            "lastnames": response.last_name,
            "admin": response.admin
        }
    
    def register(self, payload: dict) ->dict:
        user_model = UsersModel()
        user_model.insert_data(self.session, payload)

    def get_users(self, payload: dict) ->dict:
        user_model = UsersModel()

        if "role" not in payload or payload["role"] == "0":
            users = user_model.get_data_diff(self.session)
        else:
            users = user_model.get_data(self.session, payload, "all")
     
        if not users:
            raise NotFoundError("No existen usuarios.")
        return users

    def registerRest(self, payload: dict) ->dict:
        user_model = UsersModel()
        obvservability_model = ObservabilityModel()
        
        id_rest = payload["id_rest"]
        id_not_rest = payload["id_not_rest"]
        
        payload = {}
        
        ids = [id_rest, id_not_rest]
        
        response = user_model.get_data_in(ids, self.session)
        
        if not response:
            raise NotFoundError("No existen usuarios.")
        
        for item in response:
            if item.id == id_rest:
                payload["name_rest"] = item.name
                payload["last_name_rest"] = item.last_name
            
            else:
                payload["name_not_rest"] = item.name
                payload["last_name_not_rest"] = item.last_name
                
        obvservability_model.insert_data(self.session, payload)
    
    def generateExcel(self) ->dict:
        obvservability_model = ObservabilityModel()
        output = BytesIO()
        
        wb = Workbook()
        ws = wb.active

        fill1 = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        fill2 = PatternFill(start_color='ddf2ff', end_color='ddf2ff', fill_type='solid')
        
        ws["A1"] = "Persona que descansa"
        ws["B1"] = "Persona que cubre"
        ws["C1"] = "Hora de inicio"
        ws["D1"] = "Hora final"
        ws["A1"].fill = fill1
        ws["B1"].fill = fill1
        ws["C1"].fill = fill1
        ws["D1"].fill = fill1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
                
        thin_border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
        
        response = obvservability_model.get_data(self.session, response_type= "all")
        
        if not response:
            raise NotFoundError("No existen registros para exportar.")
        
        cont = 2
        
        for item in response:
            date = datetime.strptime(str(item.created_at), '%Y-%m-%d %H:%M:%S')
            
            ws.append([
                f"{item.name_rest} {item.last_name_rest}", 
                f"{item.name_not_rest} {item.last_name_not_rest}",
                str(item.created_at),
                str(date + timedelta(hours=1))
            ])
            
            ws[f"A{cont}"].fill = fill2
            ws[f"B{cont}"].fill = fill2
            ws[f"C{cont}"].fill = fill2
            ws[f"D{cont}"].fill = fill2
            
            cont += 1
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = thin_border
                
        wb.save(output)
        output.seek(0)
        
        excel_base64 = base64.b64encode(output.read()).decode('utf-8')
        
        return excel_base64